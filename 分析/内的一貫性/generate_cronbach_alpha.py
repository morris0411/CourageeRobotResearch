from __future__ import annotations

import csv
import math
from pathlib import Path
from statistics import variance

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
ANALYSIS_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = ANALYSIS_DIR / "内的一貫性"
MAPPING_PATH = ANALYSIS_DIR / "column_mapping_clean_data.csv"
EXCEL_PATH = next(ROOT.rglob("きれいデータ.xlsx"))

RESULT_CSV_PATH = OUTPUT_DIR / "cronbach_alpha_results.csv"
RESULT_MD_PATH = OUTPUT_DIR / "cronbach_alpha_results.md"
POOLED_RESULT_CSV_PATH = OUTPUT_DIR / "cronbach_alpha_results_pooled.csv"
POOLED_RESULT_MD_PATH = OUTPUT_DIR / "cronbach_alpha_results_pooled.md"

TARGET_SCALES = {"主観の勇気評定", "勇気尺度", "CZO尺度", "葛藤尺度"}


def load_mapping() -> list[dict[str, str]]:
    with MAPPING_PATH.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def load_sheet_rows() -> list[dict[str, object]]:
    workbook = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {get_column_letter(idx): value for idx, value in enumerate(row, start=1)}
        rows.append(row_data)
    return rows


def build_scale_groups(mapping_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str, str, str], list[dict[str, str]]] = {}

    for row in mapping_rows:
        scale = row["scale"]
        if scale not in TARGET_SCALES:
            continue
        key = (row["section"], row["timing"], row["video_no"], scale)
        grouped.setdefault(key, []).append(row)

    groups: list[dict[str, object]] = []
    for key, items in grouped.items():
        sorted_items = sorted(items, key=lambda r: int(r["excel_col_index"]))
        section, timing, video_no, scale = key
        groups.append(
            {
                "section": section,
                "timing": timing,
                "video_no": video_no,
                "scale": scale,
                "items": sorted_items,
            }
        )

    groups.sort(key=lambda g: (g["timing"], int(g["video_no"] or 0), str(g["scale"])))
    return groups


def reverse_score(value: float) -> float:
    return 6 - value


def cronbach_alpha(matrix: list[list[float]]) -> float:
    if not matrix or len(matrix) < 2:
        return math.nan

    item_count = len(matrix[0])
    if item_count < 2:
        return math.nan

    item_variances = []
    for col_idx in range(item_count):
        column = [row[col_idx] for row in matrix]
        item_variances.append(variance(column))

    total_scores = [sum(row) for row in matrix]
    total_variance = variance(total_scores)
    if total_variance == 0:
        return math.nan

    return (item_count / (item_count - 1)) * (1 - sum(item_variances) / total_variance)


def collect_complete_cases(
    raw_rows: list[dict[str, object]], group: dict[str, object]
) -> tuple[list[list[float]], int]:
    matrix: list[list[float]] = []

    for raw_row in raw_rows:
        scored_row: list[float] = []
        valid = True
        for item in group["items"]:
            cell_value = raw_row[item["excel_col_letter"]]
            if cell_value is None or not isinstance(cell_value, (int, float)):
                valid = False
                break
            value = float(cell_value)
            if item["reverse_scored"] == "yes":
                value = reverse_score(value)
            scored_row.append(value)

        if valid:
            matrix.append(scored_row)

    return matrix, len(matrix)


def format_alpha(alpha: float) -> str:
    if math.isnan(alpha):
        return ""
    return f"{alpha:.3f}"


def calculate_group_result(
    raw_rows: list[dict[str, object]], group: dict[str, object]
) -> dict[str, object]:
    matrix, n_valid = collect_complete_cases(raw_rows, group)
    alpha = cronbach_alpha(matrix)
    reverse_items = [item["excel_col_letter"] for item in group["items"] if item["reverse_scored"] == "yes"]

    return {
        "section": group["section"],
        "timing": group["timing"],
        "video_no": group["video_no"],
        "scale": group["scale"],
        "n_items": len(group["items"]),
        "n_valid": n_valid,
        "cronbach_alpha": format_alpha(alpha),
        "pooled_videos": "",
        "item_columns": ", ".join(item["excel_col_letter"] for item in group["items"]),
        "item_numbers": ", ".join(item["item_no_within_scale"] for item in group["items"]),
        "reverse_scored_columns": ", ".join(reverse_items),
        "reverse_scored_item_numbers": ", ".join(
            item["item_no_within_scale"] for item in group["items"] if item["reverse_scored"] == "yes"
        ),
        "reverse_rule": "1↔5, 2↔4, 3→3" if reverse_items else "",
    }


def build_pooled_post_groups(scale_groups: list[dict[str, object]]) -> list[dict[str, object]]:
    pooled: dict[str, dict[str, object]] = {}

    for group in scale_groups:
        if group["timing"] != "事後":
            continue
        scale = str(group["scale"])
        pooled.setdefault(
            scale,
            {
                "section": "事後全条件プール",
                "timing": "事後",
                "video_no": "all",
                "scale": scale,
                "items_by_video": [],
            },
        )
        pooled[scale]["items_by_video"].append(group["items"])

    pooled_groups: list[dict[str, object]] = []
    for scale, pooled_group in pooled.items():
        pooled_groups.append(pooled_group)

    pooled_groups.sort(key=lambda g: str(g["scale"]))
    return pooled_groups


def collect_pooled_complete_cases(
    raw_rows: list[dict[str, object]], pooled_group: dict[str, object]
) -> tuple[list[list[float]], int]:
    matrix: list[list[float]] = []

    for raw_row in raw_rows:
        for video_items in pooled_group["items_by_video"]:
            scored_row: list[float] = []
            valid = True
            for item in video_items:
                cell_value = raw_row[item["excel_col_letter"]]
                if cell_value is None or not isinstance(cell_value, (int, float)):
                    valid = False
                    break
                value = float(cell_value)
                if item["reverse_scored"] == "yes":
                    value = reverse_score(value)
                scored_row.append(value)

            if valid:
                matrix.append(scored_row)

    return matrix, len(matrix)


def calculate_pooled_result(
    raw_rows: list[dict[str, object]], pooled_group: dict[str, object]
) -> dict[str, object]:
    matrix, n_valid = collect_pooled_complete_cases(raw_rows, pooled_group)
    alpha = cronbach_alpha(matrix)

    first_items = pooled_group["items_by_video"][0]
    reverse_items = [item["item_no_within_scale"] for item in first_items if item["reverse_scored"] == "yes"]

    return {
        "section": pooled_group["section"],
        "timing": pooled_group["timing"],
        "video_no": pooled_group["video_no"],
        "scale": pooled_group["scale"],
        "n_items": len(first_items),
        "n_valid": n_valid,
        "cronbach_alpha": format_alpha(alpha),
        "pooled_videos": len(pooled_group["items_by_video"]),
        "item_columns": "",
        "item_numbers": ", ".join(item["item_no_within_scale"] for item in first_items),
        "reverse_scored_columns": "",
        "reverse_scored_item_numbers": ", ".join(reverse_items),
        "reverse_rule": "1↔5, 2↔4, 3→3" if reverse_items else "",
    }


def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)

    mapping_rows = load_mapping()
    raw_rows = load_sheet_rows()
    scale_groups = build_scale_groups(mapping_rows)
    pooled_post_groups = build_pooled_post_groups(scale_groups)

    results: list[dict[str, object]] = []
    for group in scale_groups:
        results.append(calculate_group_result(raw_rows, group))

    pooled_results: list[dict[str, object]] = []
    for group in scale_groups:
        if group["timing"] == "事前":
            pooled_results.append(calculate_group_result(raw_rows, group))
    for pooled_group in pooled_post_groups:
        pooled_results.append(calculate_pooled_result(raw_rows, pooled_group))

    with RESULT_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)

    with POOLED_RESULT_CSV_PATH.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(pooled_results[0].keys()))
        writer.writeheader()
        writer.writerows(pooled_results)

    md_lines = [
        "# クロンバックのα係数",
        "",
        f"- 元データ: `データ/{EXCEL_PATH.name}`",
        "- 対象シート: `きれい`",
        "- 欠損処理: 尺度ごとに完全回答者のみで算出",
        "- CZO尺度: 指定された5項目を逆転処理後に算出",
        "",
        "## 結果一覧",
        "",
        "| section | video_no | scale | n_items | n_valid | alpha | reverse_scored_columns |",
        "| --- | --- | --- | ---: | ---: | ---: | --- |",
    ]

    for result in results:
        video_no = result["video_no"] or "-"
        alpha_text = result["cronbach_alpha"] or "NA"
        reverse_cols = result["reverse_scored_columns"] or "-"
        md_lines.append(
            f"| {result['section']} | {video_no} | {result['scale']} | "
            f"{result['n_items']} | {result['n_valid']} | {alpha_text} | {reverse_cols} |"
        )

    md_lines.extend(
        [
            "",
            "## 出力ファイル",
            "",
            "- `cronbach_alpha_results.csv`: 分析用の詳細結果",
            "- `cronbach_alpha_results.md`: 読みやすい要約表",
            "",
        ]
    )

    RESULT_MD_PATH.write_text("\n".join(md_lines), encoding="utf-8")

    pooled_md_lines = [
        "# クロンバックのα係数 事前と事後全条件プール",
        "",
        f"- 元データ: `データ/{EXCEL_PATH.name}`",
        "- 対象シート: `きれい`",
        "- 事前尺度はそのまま算出",
        "- 事後尺度は4動画分をまとめてプールして算出",
        "- 欠損処理: 尺度ごとに完全回答者のみで算出",
        "- CZO尺度: 指定された5項目を逆転処理後に算出",
        "",
        "## 結果一覧",
        "",
        "| section | video_no | scale | n_items | n_valid | alpha | 備考 |",
        "| --- | --- | --- | ---: | ---: | ---: | --- |",
    ]

    for result in pooled_results:
        alpha_text = result["cronbach_alpha"] or "NA"
        if result["section"] == "事後全条件プール":
            note = f"4動画プール; reverse={result['reverse_scored_item_numbers'] or '-'}"
        else:
            note = f"reverse={result['reverse_scored_columns'] or '-'}"
        pooled_md_lines.append(
            f"| {result['section']} | {result['video_no'] or '-'} | {result['scale']} | "
            f"{result['n_items']} | {result['n_valid']} | {alpha_text} | {note} |"
        )

    pooled_md_lines.extend(
        [
            "",
            "## 出力ファイル",
            "",
            "- `cronbach_alpha_results_pooled.csv`: 事前と事後全条件プールの詳細結果",
            "- `cronbach_alpha_results_pooled.md`: 読みやすい要約表",
            "",
        ]
    )

    POOLED_RESULT_MD_PATH.write_text("\n".join(pooled_md_lines), encoding="utf-8")

    print(RESULT_CSV_PATH)
    print(RESULT_MD_PATH)
    print(POOLED_RESULT_CSV_PATH)
    print(POOLED_RESULT_MD_PATH)


if __name__ == "__main__":
    main()
