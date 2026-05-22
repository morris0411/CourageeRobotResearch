from __future__ import annotations

import csv
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import ttest_rel
from statsmodels.stats.anova import AnovaRM


ROOT = Path(__file__).resolve().parents[3]
ANALYSIS_DIR = Path(__file__).resolve().parents[2]
OUTPUT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = next(ROOT.rglob("縺阪ｌ縺・ョ繝ｼ繧ｿ.xlsx"))
MAPPING_PATH = ANALYSIS_DIR / "column_mapping_clean_data.csv"

SUMMARY_PATH = OUTPUT_DIR / "anova_summary.md"
LONG_PATH = OUTPUT_DIR / "subjective_courage_long.csv"

VIDEO_CONDITION_MAP = {
    "1": {"conflict": "葛藤あり", "action": "行動あり"},
    "2": {"conflict": "葛藤あり", "action": "行動なし"},
    "3": {"conflict": "葛藤なし", "action": "行動あり"},
    "4": {"conflict": "葛藤なし", "action": "行動なし"},
}

ITEM_TARGETS = {
    "1": "項目1",
    "2": "項目2",
    "3": "項目3",
}


def configure_plot_style() -> None:
    plt.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "MS Gothic", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False


def load_mapping() -> list[dict[str, str]]:
    with MAPPING_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_sheet() -> pd.DataFrame:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]
    df = pd.DataFrame(rows, columns=columns)
    df.insert(0, "participant_id", range(1, len(df) + 1))
    return df


def get_item_columns(mapping_rows: list[dict[str, str]], item_no: str) -> list[dict[str, str]]:
    rows = [
        row
        for row in mapping_rows
        if row["scale"] == "荳ｻ隕ｳ縺ｮ蜍・ｰ苓ｩ募ｮ・ and row["timing"] == "莠句ｾ・ and row["item_no_within_scale"] == item_no
    ]
    return sorted(rows, key=lambda r: int(r["video_no"]))


def build_long_data(df: pd.DataFrame, mapping_rows: list[dict[str, str]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for item_no, target_name in ITEM_TARGETS.items():
        item_rows = get_item_columns(mapping_rows, item_no)
        for row in item_rows:
            video_no = row["video_no"]
            tmp = pd.DataFrame(
                {
                    "participant_id": df["participant_id"],
                    "target_name": target_name,
                    "item_no": item_no,
                    "item_text": row["item_text"],
                    "video_no": video_no,
                    "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                    "action": VIDEO_CONDITION_MAP[video_no]["action"],
                    "score": pd.to_numeric(df[row["excel_col_letter"]], errors="coerce"),
                }
            )
            frames.append(tmp)

    long_df = pd.concat(frames, ignore_index=True)
    long_df["conflict"] = pd.Categorical(long_df["conflict"], categories=["葛藤なし", "葛藤あり"], ordered=True)
    long_df["action"] = pd.Categorical(long_df["action"], categories=["陦悟虚縺ゅｊ", "陦悟虚縺ｪ縺・], ordered=True)
    return long_df


def run_anova(item_df: pd.DataFrame) -> pd.DataFrame:
    analysis_df = item_df[["participant_id", "conflict", "action", "score"]].dropna().copy()
    result = AnovaRM(
        data=analysis_df,
        depvar="score",
        subject="participant_id",
        within=["conflict", "action"],
    ).fit()
    table = result.anova_table.reset_index().rename(columns={"index": "effect"})
    table["partial_eta_sq"] = (
        table["F Value"] * table["Num DF"] / (table["F Value"] * table["Num DF"] + table["Den DF"])
    )
    return table


def p_to_stars(p_value: float) -> str:
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    if p_value < 0.1:
        return "窶"
    return "ns"


def cohens_d_paired(x: pd.Series, y: pd.Series) -> float:
    diff = x - y
    return float(diff.mean() / diff.std(ddof=1))


def run_simple_effects(item_df: pd.DataFrame, interaction_p: float) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if interaction_p >= 0.1:
        return pd.DataFrame(columns=["comparison_type", "level", "group1", "group2", "t_value", "p_value", "cohens_d", "significance"])

    for conflict_level in ["闡幄陸縺ゅｊ", "闡幄陸縺ｪ縺・]:
        subset = item_df[item_df["conflict"] == conflict_level].pivot(
            index="participant_id", columns="action", values="score"
        ).dropna()
        stat, p_value = ttest_rel(subset["陦悟虚縺ゅｊ"], subset["陦悟虚縺ｪ縺・])
        d_value = cohens_d_paired(subset["陦悟虚縺ゅｊ"], subset["陦悟虚縺ｪ縺・])
        rows.append(
            {
                "comparison_type": "陦悟虚縺ｮ蜊倡ｴ比ｸｻ蜉ｹ譫・,
                "level": conflict_level,
                "group1": "陦悟虚縺ゅｊ",
                "group2": "陦悟虚縺ｪ縺・,
                "t_value": stat,
                "p_value": p_value,
                "cohens_d": d_value,
                "significance": p_to_stars(p_value),
            }
        )

    for action_level in ["陦悟虚縺ゅｊ", "陦悟虚縺ｪ縺・]:
        subset = item_df[item_df["action"] == action_level].pivot(
            index="participant_id", columns="conflict", values="score"
        ).dropna()
        stat, p_value = ttest_rel(subset["闡幄陸縺ゅｊ"], subset["闡幄陸縺ｪ縺・])
        d_value = cohens_d_paired(subset["闡幄陸縺ゅｊ"], subset["闡幄陸縺ｪ縺・])
        rows.append(
            {
                "comparison_type": "闡幄陸縺ｮ蜊倡ｴ比ｸｻ蜉ｹ譫・,
                "level": action_level,
                "group1": "闡幄陸縺ゅｊ",
                "group2": "闡幄陸縺ｪ縺・,
                "t_value": stat,
                "p_value": p_value,
                "cohens_d": d_value,
                "significance": p_to_stars(p_value),
            }
        )

    return pd.DataFrame(rows)


def annotate_simple_effects(ax: plt.Axes, means_df: pd.DataFrame, simple_df: pd.DataFrame) -> None:
    if simple_df.empty:
        return

    x_positions = {"葛藤なし": 0, "葛藤あり": 1}
    ylim_top = ax.get_ylim()[1]
    ylim_bottom = ax.get_ylim()[0]
    y_span = ylim_top - ylim_bottom
    visible_idx = 0

    for row in simple_df.itertuples(index=False):
        if row.comparison_type != "陦悟虚縺ｮ蜊倡ｴ比ｸｻ蜉ｹ譫・ or row.p_value >= 0.1:
            continue
        x = x_positions[row.level]
        high = means_df.loc[(means_df["conflict"] == row.level), "mean_score"].max()
        y = high + y_span * (0.05 + 0.06 * visible_idx)
        visible_idx += 1
        ax.plot([x - 0.08, x - 0.08, x + 0.08, x + 0.08], [y - 0.02 * y_span, y, y, y - 0.02 * y_span], color="black", lw=1)
        ax.text(x, y + 0.015 * y_span, row.significance, ha="center", va="bottom", fontsize=12)


def create_interaction_plot(item_df: pd.DataFrame, anova_df: pd.DataFrame, simple_df: pd.DataFrame, output_path: Path, y_label: str) -> None:
    means_df = (
        item_df.groupby(["conflict", "action"], observed=True)["score"]
        .mean()
        .reset_index(name="mean_score")
    )

    fig, ax = plt.subplots(figsize=(8.5, 5.2))
    x_order = ["葛藤なし", "葛藤あり"]
    x = np.arange(len(x_order))
    colors = {"陦悟虚縺ゅｊ": "#1f4e79", "陦悟虚縺ｪ縺・: "#b03a2e"}

    for action_level in ["陦悟虚縺ゅｊ", "陦悟虚縺ｪ縺・]:
        subset = means_df[means_df["action"] == action_level].set_index("conflict").loc[x_order].reset_index()
        ax.plot(x, subset["mean_score"], marker="o", linewidth=2.2, color=colors[action_level])
        y_end = subset["mean_score"].iloc[-1]
        offset = 0.12 if action_level == "陦悟虚縺ゅｊ" else -0.12
        ax.text(x[-1] + 0.08, y_end + offset, action_level, color=colors[action_level], fontsize=11, va="center")

    ax.set_xticks(x)
    ax.set_xticklabels(x_order)
    ax.set_ylabel(y_label)
    ax.set_xlabel("闡幄陸")
    ax.set_ylim(1, 7)
    ax.set_xlim(-0.15, 1.55)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(False)

    text_lines = []
    for effect_label, jp_label in [("conflict", "闡幄陸"), ("action", "陦悟虚"), ("conflict:action", "莠､莠剃ｽ懃畑")]:
        row = anova_df.loc[anova_df["effect"] == effect_label].iloc[0]
        text_lines.append(f"{jp_label}: F={row['F Value']:.3f}, p={row['Pr > F']:.3f}")
    ax.text(
        0.02,
        0.98,
        "\n".join(text_lines),
        transform=ax.transAxes,
        ha="left",
        va="top",
        fontsize=10,
        bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": "none", "alpha": 0.9},
    )

    interaction_p = float(anova_df.loc[anova_df["effect"] == "conflict:action", "Pr > F"].iloc[0])
    if interaction_p < 0.1:
        annotate_simple_effects(ax, means_df, simple_df)

    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def main() -> None:
    configure_plot_style()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    mapping_rows = load_mapping()
    df = load_sheet()
    long_df = build_long_data(df, mapping_rows)
    long_df.to_csv(LONG_PATH, index=False, encoding="utf-8-sig")

    summary_lines = [
        "# 荳ｻ隕ｳ蜍・ｰ苓ｩ募ｮ壹・2隕∝屏蛻・淵蛻・梵",
        "",
        f"- 蜈・ョ繝ｼ繧ｿ: `繝・・繧ｿ/{EXCEL_PATH.name}`",
        "- 蛻・梵蟇ｾ雎｡: 荳ｻ隕ｳ蜍・ｰ苓ｩ募ｮ・鬆・岼",
        "- 蛻・梵豕・ 蜿ょ刈閠・・ 2隕∝屏蛻・淵蛻・梵・・novaRM・・,
        "- 隕∝屏1: 闡幄陸・郁騒阯､縺ゅｊ / 闡幄陸縺ｪ縺暦ｼ・,
        "- 隕∝屏2: 陦悟虚・郁｡悟虚縺ゅｊ / 陦悟虚縺ｪ縺暦ｼ・,
        "- 蜍慕判譚｡莉ｶ縺ｮ蟇ｾ蠢應ｻｮ螳・ 蜍慕判1=闡幄陸縺ゅｊﾃ苓｡悟虚縺ゅｊ, 蜍慕判2=闡幄陸縺ゅｊﾃ苓｡悟虚縺ｪ縺・ 蜍慕判3=闡幄陸縺ｪ縺療苓｡悟虚縺ゅｊ, 蜍慕判4=闡幄陸縺ｪ縺療苓｡悟虚縺ｪ縺・,
        "- 荳ｻ隕ｳ蜍・ｰ苓ｩ募ｮ壹・蟆ｺ蠎ｦ蛹悶○縺壹∝推鬆・岼繧貞句挨縺ｫ蛻・梵",
        "- 莠句燕蛟､縺後↑縺・◆繧√∝ｾ灘ｱ槫､画焚縺ｯ蜷・虚逕ｻ隕冶・蠕後せ繧ｳ繧｢縺ｮ縺ｿ",
        "- 繧ｰ繝ｩ繝穂ｻ墓ｧ・ `繝励Ο繝・ヨ蝗ｳ繝ｫ繝ｼ繝ｫ.txt` 縺ｫ貅匁侠",
        "",
        "## 蜃ｺ蜉帙ヵ繧｡繧､繝ｫ",
        "",
        "- `subjective_courage_long.csv`: 蛻・梵逕ｨ繝ｭ繝ｳ繧ｰ繝・・繧ｿ",
    ]

    for item_no, target_name in ITEM_TARGETS.items():
        item_df = long_df[long_df["item_no"] == item_no].copy()
        anova_df = run_anova(item_df)
        interaction_p = float(anova_df.loc[anova_df["effect"] == "conflict:action", "Pr > F"].iloc[0])
        simple_df = run_simple_effects(item_df, interaction_p)

        anova_path = OUTPUT_DIR / f"anova_item{item_no}.csv"
        simple_path = OUTPUT_DIR / f"simple_effects_item{item_no}.csv"
        plot_path = OUTPUT_DIR / f"anova_plot_item{item_no}.png"

        anova_df.to_csv(anova_path, index=False, encoding="utf-8-sig")
        simple_df.to_csv(simple_path, index=False, encoding="utf-8-sig")
        create_interaction_plot(item_df, anova_df, simple_df, plot_path, f"荳ｻ隕ｳ蜍・ｰ苓ｩ募ｮ・鬆・岼{item_no}")

        means_df = (
            item_df.groupby(["conflict", "action"], observed=True)["score"]
            .mean()
            .reset_index(name="mean_score")
        )

        summary_lines.extend(
            [
                f"- `anova_item{item_no}.csv`: 鬆・岼{item_no}縺ｮANOVA陦ｨ",
                f"- `simple_effects_item{item_no}.csv`: 鬆・岼{item_no}縺ｮ蜊倡ｴ比ｸｻ蜉ｹ譫・,
                f"- `anova_plot_item{item_no}.png`: 鬆・岼{item_no}縺ｮ莠､莠剃ｽ懃畑繝励Ο繝・ヨ",
                "",
                f"## 鬆・岼{item_no}",
                "",
                f"- 鬆・岼譁・ {ITEM_TARGETS[item_no].split('_', 1)[1]}",
                "",
                "| conflict | action | mean |",
                "| --- | --- | ---: |",
            ]
        )

        for row in means_df.itertuples(index=False):
            summary_lines.append(f"| {row.conflict} | {row.action} | {row.mean_score:.4f} |")

        summary_lines.extend(
            [
                "",
                "| effect | F | Num DF | Den DF | p | partial ﾎｷﾂｲ |",
                "| --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )

        for _, row in anova_df.iterrows():
            summary_lines.append(
                f"| {row['effect']} | {row['F Value']:.6f} | {row['Num DF']:.1f} | {row['Den DF']:.1f} | {row['Pr > F']:.6f} | {row['partial_eta_sq']:.6f} |"
            )

        if simple_df.empty:
            summary_lines.extend(["", "- 莠､莠剃ｽ懃畑縺梧怏諢上〒縺ｪ縺・◆繧∝腰邏比ｸｻ蜉ｹ譫懊・螳滓命縺帙★", ""])
        else:
            summary_lines.extend(["", simple_df.to_markdown(index=False), ""])

    SUMMARY_PATH.write_text("\n".join(summary_lines), encoding="utf-8")

    print(LONG_PATH)
    print(SUMMARY_PATH)
    for item_no in ITEM_TARGETS:
        print(OUTPUT_DIR / f"anova_item{item_no}.csv")
        print(OUTPUT_DIR / f"simple_effects_item{item_no}.csv")
        print(OUTPUT_DIR / f"anova_plot_item{item_no}.png")


if __name__ == "__main__":
    main()


