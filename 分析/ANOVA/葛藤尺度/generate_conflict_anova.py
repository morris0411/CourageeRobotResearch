from __future__ import annotations

import csv
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr, ttest_rel
from statsmodels.stats.anova import AnovaRM


ROOT = Path(__file__).resolve().parents[3]
ANALYSIS_DIR = Path(__file__).resolve().parents[2]
OUTPUT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = next(ROOT.rglob("きれいデータ.xlsx"))
MAPPING_PATH = ANALYSIS_DIR / "column_mapping_clean_data.csv"

LONG_PATH = OUTPUT_DIR / "conflict_scores_long.csv"
ANOVA_PATH = OUTPUT_DIR / "anova_post_score.csv"
SIMPLE_PATH = OUTPUT_DIR / "simple_effects_post_score.csv"
CORR_PATH = OUTPUT_DIR / "item5_correlation.csv"
SUMMARY_PATH = OUTPUT_DIR / "anova_summary.md"
PLOT_PATH = OUTPUT_DIR / "anova_plot_post_score.png"

VIDEO_CONDITION_MAP = {
    "1": {"conflict": "葛藤あり", "action": "行動あり"},
    "2": {"conflict": "葛藤あり", "action": "行動なし"},
    "3": {"conflict": "葛藤なし", "action": "行動あり"},
    "4": {"conflict": "葛藤なし", "action": "行動なし"},
}
CONFLICT_ORDER = ["葛藤なし", "葛藤あり"]
ACTION_ORDER = ["行動あり", "行動なし"]
COLORS = {"行動あり": "#1f4e79", "行動なし": "#b03a2e"}


def configure_plot_style() -> None:
    plt.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "MS Gothic", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False


def load_mapping() -> list[dict[str, str]]:
    with MAPPING_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_sheet() -> pd.DataFrame:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]
    df = pd.DataFrame(rows, columns=columns)
    df.insert(0, "participant_id", range(1, len(df) + 1))
    return df


def get_scale_columns(mapping_rows: list[dict[str, str]], scale: str, video_no: str) -> list[dict[str, str]]:
    rows = [
        row for row in mapping_rows if row["scale"] == scale and row["timing"] == "事後" and row["video_no"] == video_no
    ]
    return sorted(rows, key=lambda r: int(r["item_no_within_scale"]))


def score_from_columns(df: pd.DataFrame, items: list[dict[str, str]]) -> pd.Series:
    cols = [pd.to_numeric(df[item["excel_col_letter"]], errors="coerce") for item in items]
    return pd.concat(cols, axis=1).mean(axis=1, skipna=False)


def build_long_data(df: pd.DataFrame, mapping_rows: list[dict[str, str]]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for video_no in ["1", "2", "3", "4"]:
        scale_items = get_scale_columns(mapping_rows, "葛藤尺度", video_no)
        item5 = get_scale_columns(mapping_rows, "葛藤確認項目", video_no)[0]
        scale_score = score_from_columns(df, scale_items)
        item5_score = pd.to_numeric(df[item5["excel_col_letter"]], errors="coerce")
        frames.append(
            pd.DataFrame(
                {
                    "participant_id": df["participant_id"],
                    "video_no": video_no,
                    "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                    "action": VIDEO_CONDITION_MAP[video_no]["action"],
                    "scale_score": scale_score,
                    "item5_score": item5_score,
                }
            )
        )
    long_df = pd.concat(frames, ignore_index=True)
    long_df["conflict"] = pd.Categorical(long_df["conflict"], categories=CONFLICT_ORDER, ordered=True)
    long_df["action"] = pd.Categorical(long_df["action"], categories=ACTION_ORDER, ordered=True)
    return long_df


def run_anova(long_df: pd.DataFrame) -> pd.DataFrame:
    analysis_df = long_df[["participant_id", "conflict", "action", "scale_score"]].dropna().copy()
    result = AnovaRM(data=analysis_df, depvar="scale_score", subject="participant_id", within=["conflict", "action"]).fit()
    table = result.anova_table.reset_index().rename(columns={"index": "effect"})
    table["partial_eta_sq"] = table["F Value"] * table["Num DF"] / (table["F Value"] * table["Num DF"] + table["Den DF"])
    return table


def p_to_stars(p_value: float) -> str:
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    if p_value < 0.1:
        return "†"
    return "ns"


def cohens_d_paired(x: pd.Series, y: pd.Series) -> float:
    diff = x - y
    return float(diff.mean() / diff.std(ddof=1))


def run_simple_effects(long_df: pd.DataFrame, interaction_p: float) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if interaction_p >= 0.1:
        return pd.DataFrame(columns=["comparison_type", "level", "group1", "group2", "t_value", "p_value", "cohens_d", "significance"])

    for conflict_level in CONFLICT_ORDER:
        subset = long_df[long_df["conflict"] == conflict_level].pivot(index="participant_id", columns="action", values="scale_score").dropna()
        stat, p_value = ttest_rel(subset["行動あり"], subset["行動なし"])
        rows.append({
            "comparison_type": "行動の単純主効果",
            "level": conflict_level,
            "group1": "行動あり",
            "group2": "行動なし",
            "t_value": stat,
            "p_value": p_value,
            "cohens_d": cohens_d_paired(subset["行動あり"], subset["行動なし"]),
            "significance": p_to_stars(float(p_value)),
        })

    for action_level in ACTION_ORDER:
        subset = long_df[long_df["action"] == action_level].pivot(index="participant_id", columns="conflict", values="scale_score").dropna()
        stat, p_value = ttest_rel(subset["葛藤あり"], subset["葛藤なし"])
        rows.append({
            "comparison_type": "葛藤の単純主効果",
            "level": action_level,
            "group1": "葛藤あり",
            "group2": "葛藤なし",
            "t_value": stat,
            "p_value": p_value,
            "cohens_d": cohens_d_paired(subset["葛藤あり"], subset["葛藤なし"]),
            "significance": p_to_stars(float(p_value)),
        })

    return pd.DataFrame(rows)


def run_correlations(long_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for video_no in ["1", "2", "3", "4"]:
        subset = long_df[long_df["video_no"] == video_no][["scale_score", "item5_score"]].dropna()
        r_value, p_value = pearsonr(subset["scale_score"], subset["item5_score"])
        rows.append({"scope": f"動画{video_no}", "video_no": video_no, "n": int(len(subset)), "pearson_r": r_value, "p_value": p_value})
    pooled = long_df[["scale_score", "item5_score"]].dropna()
    r_value, p_value = pearsonr(pooled["scale_score"], pooled["item5_score"])
    rows.append({"scope": "全動画プール", "video_no": "all", "n": int(len(pooled)), "pearson_r": r_value, "p_value": p_value})
    return pd.DataFrame(rows)


def annotate_simple_effects(ax: plt.Axes, means_df: pd.DataFrame, simple_df: pd.DataFrame) -> None:
    if simple_df.empty:
        return
    x_positions = {label: idx for idx, label in enumerate(CONFLICT_ORDER)}
    ylim_bottom, ylim_top = ax.get_ylim()
    y_span = ylim_top - ylim_bottom
    visible_idx = 0
    for row in simple_df.itertuples(index=False):
        if row.comparison_type != "行動の単純主効果" or float(row.p_value) >= 0.1:
            continue
        x = x_positions[row.level]
        high = means_df.loc[means_df["conflict"] == row.level, "mean_score"].max()
        y = high + y_span * (0.05 + 0.06 * visible_idx)
        visible_idx += 1
        ax.plot([x - 0.08, x - 0.08, x + 0.08, x + 0.08], [y - 0.02 * y_span, y, y, y - 0.02 * y_span], color="black", lw=1)
        ax.text(x, y + 0.015 * y_span, row.significance, ha="center", va="bottom", fontsize=12)


def create_interaction_plot(long_df: pd.DataFrame, anova_df: pd.DataFrame, simple_df: pd.DataFrame, output_path: Path) -> None:
    means_df = long_df.groupby(["conflict", "action"], observed=True)["scale_score"].mean().reset_index(name="mean_score")
    fig, ax = plt.subplots(figsize=(8.5, 5.2))
    x = np.arange(len(CONFLICT_ORDER))
    for action_level in ACTION_ORDER:
        subset = means_df[means_df["action"] == action_level].set_index("conflict").loc[CONFLICT_ORDER].reset_index()
        ax.plot(x, subset["mean_score"], marker="o", linewidth=2.2, color=COLORS[action_level])
        y_end = subset["mean_score"].iloc[-1]
        offset = 0.12 if action_level == "行動あり" else -0.12
        ax.text(x[-1] + 0.08, y_end + offset, action_level, color=COLORS[action_level], fontsize=11, va="center")
    ax.set_xticks(x)
    ax.set_xticklabels(CONFLICT_ORDER)
    ax.set_ylabel("葛藤尺度得点")
    ax.set_xlabel("葛藤")
    ax.set_ylim(1, 7)
    ax.set_xlim(-0.15, 1.55)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(False)
    text_lines = []
    for effect_label, jp_label in [("conflict", "葛藤"), ("action", "行動"), ("conflict:action", "交互作用")]:
        row = anova_df.loc[anova_df["effect"] == effect_label].iloc[0]
        text_lines.append(f"{jp_label}: F={float(row['F Value']):.3f}, p={float(row['Pr > F']):.3f}")
    ax.text(0.02, 0.98, "\n".join(text_lines), transform=ax.transAxes, ha="left", va="top", fontsize=10, bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": "none", "alpha": 0.9})
    if float(anova_df.loc[anova_df["effect"] == "conflict:action", "Pr > F"].iloc[0]) < 0.1:
        annotate_simple_effects(ax, means_df, simple_df)
    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def write_summary(long_df: pd.DataFrame, anova_df: pd.DataFrame, simple_df: pd.DataFrame, corr_df: pd.DataFrame) -> None:
    mean_table = long_df.groupby(["conflict", "action"], observed=True)[["scale_score", "item5_score"]].mean().reset_index()
    lines = [
        "# 葛藤尺度の2要因分散分析",
        "",
        f"- 元データ: `データ/{EXCEL_PATH.name}`",
        "",
        "## 条件平均",
        "",
        "| conflict | action | scale_score_mean | item5_mean |",
        "| --- | --- | ---: | ---: |",
    ]
    for row in mean_table.itertuples(index=False):
        lines.append(f"| {row.conflict} | {row.action} | {row.scale_score:.4f} | {row.item5_score:.4f} |")
    lines.extend(["", "## ANOVA", "", "| effect | F | Num DF | Den DF | p | partial η² |", "| --- | ---: | ---: | ---: | ---: | ---: |"])
    for _, row in anova_df.iterrows():
        lines.append(f"| {row['effect']} | {row['F Value']:.6f} | {row['Num DF']:.1f} | {row['Den DF']:.1f} | {row['Pr > F']:.6f} | {row['partial_eta_sq']:.6f} |")
    lines.extend(["", "## 単純主効果", "", simple_df.to_markdown(index=False) if not simple_df.empty else "- 交互作用が有意でないため実施せず"])
    lines.extend(["", "## 相関", "", "| scope | n | r | p |", "| --- | ---: | ---: | ---: |"])
    for row in corr_df.itertuples(index=False):
        lines.append(f"| {row.scope} | {row.n} | {row.pearson_r:.6f} | {row.p_value:.6f} |")
    SUMMARY_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    configure_plot_style()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    mapping_rows = load_mapping()
    df = load_sheet()
    long_df = build_long_data(df, mapping_rows)
    anova_df = run_anova(long_df)
    simple_df = run_simple_effects(long_df, float(anova_df.loc[anova_df["effect"] == "conflict:action", "Pr > F"].iloc[0]))
    corr_df = run_correlations(long_df)
    long_df.to_csv(LONG_PATH, index=False, encoding="utf-8-sig")
    anova_df.to_csv(ANOVA_PATH, index=False, encoding="utf-8-sig")
    simple_df.to_csv(SIMPLE_PATH, index=False, encoding="utf-8-sig")
    corr_df.to_csv(CORR_PATH, index=False, encoding="utf-8-sig")
    create_interaction_plot(long_df, anova_df, simple_df, PLOT_PATH)
    write_summary(long_df, anova_df, simple_df, corr_df)


if __name__ == "__main__":
    main()
