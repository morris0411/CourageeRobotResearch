from __future__ import annotations

import csv
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from statsmodels.stats.anova import AnovaRM


ROOT = Path(__file__).resolve().parents[3]
ANALYSIS_DIR = Path(__file__).resolve().parents[2]
OUTPUT_DIR = Path(__file__).resolve().parent
EXCEL_PATH = ROOT / "データ" / "きれいデータ.xlsx"
MAPPING_PATH = ANALYSIS_DIR / "column_mapping_clean_data.csv"

VIDEO_CONDITION_MAP = {
    "1": {"conflict": "葛藤あり", "action": "行動あり"},
    "2": {"conflict": "葛藤あり", "action": "行動なし"},
    "3": {"conflict": "葛藤なし", "action": "行動あり"},
    "4": {"conflict": "葛藤なし", "action": "行動なし"},
}

STRATA = [
    {
        "name": "勇気事前3.5以下",
        "source_scale": "勇気尺度",
        "threshold": 3.5,
        "operator_label": "<=",
    },
    {
        "name": "勇気事前3.5以上",
        "source_scale": "勇気尺度",
        "threshold": 3.5,
        "operator_label": ">=",
    },
    {
        "name": "CZO事前2.5以下",
        "source_scale": "CZO尺度",
        "threshold": 2.5,
        "operator_label": "<=",
    },
    {
        "name": "CZO事前2.5以上",
        "source_scale": "CZO尺度",
        "threshold": 2.5,
        "operator_label": ">=",
    },
]

POST_SCALES = [
    {"scale": "勇気尺度", "has_pre": True},
    {"scale": "CZO尺度", "has_pre": True},
    {"scale": "葛藤尺度", "has_pre": False},
]

ITEM_ANALYSES = [
    {"scale": "主観の勇気評定", "has_pre": False},
]

ACTION_ORDER = ["行動あり", "行動なし"]
CONFLICT_ORDER = ["葛藤なし", "葛藤あり"]
ACTION_COLORS = {"行動あり": "#1f4e79", "行動なし": "#b03a2e"}


def configure_plot_style() -> None:
    plt.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "MS Gothic", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False


def load_mapping() -> list[dict[str, str]]:
    with MAPPING_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_sheet() -> pd.DataFrame:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]
    df = pd.DataFrame(rows, columns=columns)
    df.insert(0, "participant_id", range(1, len(df) + 1))
    return df


def get_mapping_rows(
    mapping_rows: list[dict[str, str]],
    *,
    scale: str,
    timing: str,
    video_no: str | None = None,
) -> list[dict[str, str]]:
    rows = [
        row
        for row in mapping_rows
        if row["scale"] == scale and row["timing"] == timing and (video_no is None or row["video_no"] == video_no)
    ]
    return sorted(rows, key=lambda r: int(r["item_no_within_scale"]))


def reverse_if_needed(series: pd.Series, row: dict[str, str]) -> pd.Series:
    if row["reverse_scored"] == "yes":
        return 6 - series
    return series


def score_items(df: pd.DataFrame, rows: list[dict[str, str]]) -> pd.Series:
    scored = []
    for row in rows:
        series = pd.to_numeric(df[row["excel_col_letter"]], errors="coerce")
        scored.append(reverse_if_needed(series, row))
    return pd.concat(scored, axis=1).mean(axis=1, skipna=False)


def get_pre_scale_score(df: pd.DataFrame, mapping_rows: list[dict[str, str]], scale: str) -> pd.Series:
    rows = get_mapping_rows(mapping_rows, scale=scale, timing="事前")
    return score_items(df, rows)


def build_scale_long(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    scale: str,
    has_pre: bool,
) -> pd.DataFrame:
    pre_score = get_pre_scale_score(df, mapping_rows, scale) if has_pre else None
    frames: list[pd.DataFrame] = []
    for video_no in ["1", "2", "3", "4"]:
        post_rows = get_mapping_rows(mapping_rows, scale=scale, timing="事後", video_no=video_no)
        post_score = score_items(df, post_rows)
        frame = pd.DataFrame(
            {
                "participant_id": df["participant_id"],
                "video_no": video_no,
                "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                "action": VIDEO_CONDITION_MAP[video_no]["action"],
                "score": post_score,
                "score_type": "post",
                "target_kind": "scale",
                "scale": scale,
                "item_no": pd.NA,
                "item_text": pd.NA,
            }
        )
        frames.append(frame)
        if has_pre and pre_score is not None:
            diff_frame = frame.copy()
            diff_frame["score"] = post_score - pre_score
            diff_frame["score_type"] = "diff"
            frames.append(diff_frame)
    return pd.concat(frames, ignore_index=True)


def build_item_long(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    scale: str,
    has_pre: bool,
) -> pd.DataFrame:
    pre_rows = {}
    if has_pre:
        pre_rows = {
            row["item_no_within_scale"]: row
            for row in get_mapping_rows(mapping_rows, scale=scale, timing="事前")
        }

    frames: list[pd.DataFrame] = []
    for video_no in ["1", "2", "3", "4"]:
        post_rows = get_mapping_rows(mapping_rows, scale=scale, timing="事後", video_no=video_no)
        for row in post_rows:
            post_series = pd.to_numeric(df[row["excel_col_letter"]], errors="coerce")
            post_series = reverse_if_needed(post_series, row)
            frame = pd.DataFrame(
                {
                    "participant_id": df["participant_id"],
                    "video_no": video_no,
                    "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                    "action": VIDEO_CONDITION_MAP[video_no]["action"],
                    "score": post_series,
                    "score_type": "post",
                    "target_kind": "item",
                    "scale": scale,
                    "item_no": row["item_no_within_scale"],
                    "item_text": row["item_text"],
                }
            )
            frames.append(frame)
            if has_pre:
                pre_row = pre_rows[row["item_no_within_scale"]]
                pre_series = pd.to_numeric(df[pre_row["excel_col_letter"]], errors="coerce")
                pre_series = reverse_if_needed(pre_series, pre_row)
                diff_frame = frame.copy()
                diff_frame["score"] = post_series - pre_series
                diff_frame["score_type"] = "diff"
                frames.append(diff_frame)
    return pd.concat(frames, ignore_index=True)


def finalize_long(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["conflict"] = pd.Categorical(df["conflict"], categories=["葛藤なし", "葛藤あり"], ordered=True)
    df["action"] = pd.Categorical(df["action"], categories=["行動あり", "行動なし"], ordered=True)
    return df


def run_anova(target_df: pd.DataFrame) -> pd.DataFrame:
    analysis_df = target_df[["participant_id", "conflict", "action", "score"]].dropna().copy()
    if analysis_df["participant_id"].nunique() < 2:
        return pd.DataFrame()
    result = AnovaRM(
        data=analysis_df,
        depvar="score",
        subject="participant_id",
        within=["conflict", "action"],
    ).fit()
    table = (
        result.anova_table.reset_index()
        .rename(
            columns={
                "index": "effect",
                "F Value": "F_value",
                "Num DF": "num_df",
                "Den DF": "den_df",
                "Pr > F": "p_value",
            }
        )
    )
    table["partial_eta_sq"] = (
        table["F_value"] * table["num_df"] / (table["F_value"] * table["num_df"] + table["den_df"])
    )
    return table


def p_to_stars(p_value: float) -> str:
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    if p_value < 0.1:
        return "†"
    return "ns"


def make_target_slug(target_kind: str, scale: str, item_no: object, score_type: str) -> str:
    scale_slug = (
        scale.replace("尺度", "")
        .replace("確認項目", "確認項目")
        .replace("主観の", "主観")
        .replace(" ", "_")
    )
    if target_kind == "scale":
        return f"{scale_slug}_{score_type}"
    return f"{scale_slug}_item{item_no}_{score_type}"


def resolve_y_limits(scale: str, score_type: str, values: pd.Series) -> tuple[float, float]:
    if score_type == "post":
        if scale == "CZO尺度":
            return (1.0, 5.0)
        return (1.0, 7.0)

    min_value = float(values.min())
    max_value = float(values.max())
    bound = max(abs(min_value), abs(max_value))
    bound = max(0.5, np.ceil((bound + 0.05) * 10) / 10)
    return (-bound, bound)


def build_anova_text(anova_df: pd.DataFrame) -> str:
    label_map = {"conflict": "葛藤", "action": "行動", "conflict:action": "交互作用"}
    lines = []
    for effect in ["conflict", "action", "conflict:action"]:
        row = anova_df.loc[anova_df["effect"] == effect].iloc[0]
        lines.append(f"{label_map[effect]}: F={row['F_value']:.3f}, p={row['p_value']:.3f}")
    return "\n".join(lines)


def annotate_significance(ax: plt.Axes, means_df: pd.DataFrame, anova_df: pd.DataFrame) -> None:
    significant = anova_df[anova_df["p_value"] < 0.1].copy()
    if significant.empty:
        return

    x_positions = {label: idx for idx, label in enumerate(CONFLICT_ORDER)}
    ylim_bottom, ylim_top = ax.get_ylim()
    y_span = ylim_top - ylim_bottom
    used_index = 0

    for effect in ["conflict", "action", "conflict:action"]:
        effect_rows = significant[significant["effect"] == effect]
        if effect_rows.empty:
            continue
        row = effect_rows.iloc[0]
        if effect == "action":
            for conflict_level in CONFLICT_ORDER:
                x = x_positions[conflict_level]
                high = means_df.loc[means_df["conflict"] == conflict_level, "mean_score"].max()
                y = high + y_span * (0.05 + 0.07 * used_index)
                ax.plot(
                    [x - 0.08, x - 0.08, x + 0.08, x + 0.08],
                    [y - 0.02 * y_span, y, y, y - 0.02 * y_span],
                    color="black",
                    lw=1,
                )
                ax.text(x, y + 0.015 * y_span, p_to_stars(float(row["p_value"])), ha="center", va="bottom", fontsize=12)
                used_index += 1
        elif effect == "conflict":
            for action_level in ACTION_ORDER:
                subset = means_df[means_df["action"] == action_level].set_index("conflict")
                x_left = x_positions["葛藤なし"]
                x_right = x_positions["葛藤あり"]
                high = subset.loc[CONFLICT_ORDER, "mean_score"].max()
                y = high + y_span * (0.05 + 0.07 * used_index)
                delta = -0.03 if action_level == "行動あり" else 0.03
                ax.plot(
                    [x_left + delta, x_left + delta, x_right + delta, x_right + delta],
                    [y - 0.02 * y_span, y, y, y - 0.02 * y_span],
                    color="black",
                    lw=1,
                )
                ax.text((x_left + x_right) / 2 + delta, y + 0.015 * y_span, p_to_stars(float(row["p_value"])), ha="center", va="bottom", fontsize=12)
                used_index += 1
        else:
            y = means_df["mean_score"].max() + y_span * (0.08 + 0.07 * used_index)
            ax.text(0.5, y, p_to_stars(float(row["p_value"])), ha="center", va="bottom", fontsize=12)
            used_index += 1


def create_interaction_plot(
    target_df: pd.DataFrame,
    mean_df: pd.DataFrame,
    anova_df: pd.DataFrame,
    output_path: Path,
) -> None:
    if anova_df.empty or mean_df.empty:
        return

    y_limits = resolve_y_limits(
        scale=str(target_df["scale"].iloc[0]),
        score_type=str(target_df["score_type"].iloc[0]),
        values=target_df["score"].dropna(),
    )
    y_label = str(target_df["scale"].iloc[0])
    if str(target_df["target_kind"].iloc[0]) == "item":
        y_label += f" 項目{target_df['item_no'].iloc[0]}"
    if str(target_df["score_type"].iloc[0]) == "diff":
        y_label += " 差分"
    else:
        y_label += " 得点"

    plot_means = (
        mean_df[["conflict", "action", "mean_score"]]
        .drop_duplicates()
        .copy()
    )

    fig, ax = plt.subplots(figsize=(8.5, 5.2))
    x = np.arange(len(CONFLICT_ORDER))
    for action_level in ACTION_ORDER:
        subset = plot_means[plot_means["action"] == action_level].set_index("conflict").loc[CONFLICT_ORDER].reset_index()
        ax.plot(x, subset["mean_score"], marker="o", linewidth=2.2, color=ACTION_COLORS[action_level])
        y_end = float(subset["mean_score"].iloc[-1])
        offset = 0.04 * (y_limits[1] - y_limits[0])
        offset = offset if action_level == "行動あり" else -offset
        ax.text(x[-1] + 0.08, y_end + offset, action_level, color=ACTION_COLORS[action_level], fontsize=11, va="center")

    ax.set_xticks(x)
    ax.set_xticklabels(CONFLICT_ORDER)
    ax.set_xlabel("葛藤")
    ax.set_ylabel(y_label)
    ax.set_xlim(-0.15, 1.55)
    ax.set_ylim(*y_limits)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(False)
    ax.text(
        0.02,
        0.98,
        build_anova_text(anova_df),
        transform=ax.transAxes,
        ha="left",
        va="top",
        fontsize=10,
        bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": "none", "alpha": 0.9},
    )
    annotate_significance(ax, plot_means, anova_df)
    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def summarize_targets(long_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    mean_rows: list[dict[str, object]] = []
    anova_rows: list[dict[str, object]] = []

    keys = ["target_kind", "scale", "item_no", "item_text", "score_type"]
    for key_values, target_df in long_df.groupby(keys, dropna=False, observed=True):
        target_kind, scale, item_no, item_text, score_type = key_values
        clean_df = target_df.dropna(subset=["score"]).copy()
        n_participants = int(clean_df["participant_id"].nunique())
        means = (
            clean_df.groupby(["conflict", "action"], observed=True)["score"]
            .mean()
            .reset_index(name="mean_score")
        )
        for row in means.itertuples(index=False):
            mean_rows.append(
                {
                    "target_kind": target_kind,
                    "scale": scale,
                    "item_no": item_no,
                    "item_text": item_text,
                    "score_type": score_type,
                    "n_participants": n_participants,
                    "conflict": row.conflict,
                    "action": row.action,
                    "mean_score": row.mean_score,
                }
            )

        anova_df = run_anova(clean_df)
        if anova_df.empty:
            continue
        for row in anova_df.itertuples(index=False):
            anova_rows.append(
                {
                    "target_kind": target_kind,
                    "scale": scale,
                    "item_no": item_no,
                    "item_text": item_text,
                    "score_type": score_type,
                    "n_participants": n_participants,
                    "effect": row.effect,
                    "F_value": row.F_value,
                    "num_df": row.num_df,
                    "den_df": row.den_df,
                    "p_value": row.p_value,
                    "partial_eta_sq": row.partial_eta_sq,
                }
            )

    return pd.DataFrame(mean_rows), pd.DataFrame(anova_rows)


def subset_mask(series: pd.Series, operator_label: str, threshold: float) -> pd.Series:
    if operator_label == "<=":
        return series <= threshold
    if operator_label == ">=":
        return series >= threshold
    raise ValueError(f"Unknown operator: {operator_label}")


def format_target_label(row: pd.Series) -> str:
    if row["target_kind"] == "scale":
        return f"{row['scale']}（{row['score_type']}）"
    return f"{row['scale']} 項目{row['item_no']}（{row['score_type']}）"


def write_summary(
    out_dir: Path,
    stratum: dict[str, object],
    group_n: int,
    overlap_n: int,
    mean_df: pd.DataFrame,
    anova_df: pd.DataFrame,
    plot_rows: list[dict[str, object]],
) -> None:
    lines = [
        f"# {stratum['name']} の層別ANOVA",
        "",
        f"- 事前尺度: {stratum['source_scale']}",
        f"- 条件: {stratum['source_scale']} 事前得点 {stratum['operator_label']} {stratum['threshold']}",
        f"- 対象者数: {group_n}",
        f"- しきい値ちょうどの人数: {overlap_n}",
        "- 注記: ユーザー指定どおり「以下」「以上」をそのまま用いたため、しきい値ちょうどの回答者は両群に含まれる。",
        "",
        "## ANOVA結果",
        "",
        "| target | effect | F | Num DF | Den DF | p | partial η² | n |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]

    if anova_df.empty:
        lines.append("| 分析可能な対象なし | - | - | - | - | - | - | - |")
    else:
        for _, row in anova_df.sort_values(
            ["target_kind", "scale", "item_no", "score_type", "effect"],
            na_position="last",
        ).iterrows():
            lines.append(
                f"| {format_target_label(row)} | {row['effect']} | {row['F_value']:.6f} | "
                f"{row['num_df']:.1f} | {row['den_df']:.1f} | {row['p_value']:.6f} | "
                f"{row['partial_eta_sq']:.6f} | {int(row['n_participants'])} |"
            )

    lines.extend(
        [
            "",
            "## 条件平均",
            "",
            "| target | conflict | action | mean | n |",
            "| --- | --- | --- | ---: | ---: |",
        ]
    )

    if mean_df.empty:
        lines.append("| 分析可能な対象なし | - | - | - | - |")
    else:
        for _, row in mean_df.sort_values(
            ["target_kind", "scale", "item_no", "score_type", "conflict", "action"],
            na_position="last",
        ).iterrows():
            lines.append(
                f"| {format_target_label(row)} | {row['conflict']} | {row['action']} | "
                f"{row['mean_score']:.6f} | {int(row['n_participants'])} |"
            )

    lines.extend(
        [
            "",
            "## プロット",
            "",
            "| target | file |",
            "| --- | --- |",
        ]
    )
    for plot_row in plot_rows:
        lines.append(f"| {plot_row['target']} | `{plot_row['file']}` |")

    (out_dir / "summary.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    configure_plot_style()

    mapping_rows = load_mapping()
    raw_df = load_sheet()

    pre_scores = {
        "勇気尺度": get_pre_scale_score(raw_df, mapping_rows, "勇気尺度"),
        "CZO尺度": get_pre_scale_score(raw_df, mapping_rows, "CZO尺度"),
    }

    all_frames = []
    for spec in POST_SCALES:
        all_frames.append(build_scale_long(raw_df, mapping_rows, spec["scale"], spec["has_pre"]))
    for spec in ITEM_ANALYSES:
        all_frames.append(build_item_long(raw_df, mapping_rows, spec["scale"], spec["has_pre"]))
    long_df = finalize_long(pd.concat(all_frames, ignore_index=True))

    stratum_index_rows = []
    for stratum in STRATA:
        out_dir = OUTPUT_DIR / stratum["name"]
        out_dir.mkdir(parents=True, exist_ok=True)
        plot_dir = out_dir / "plots"
        plot_dir.mkdir(parents=True, exist_ok=True)

        base_series = pre_scores[stratum["source_scale"]]
        mask = subset_mask(base_series, stratum["operator_label"], float(stratum["threshold"]))
        overlap = int((base_series == float(stratum["threshold"])).sum())
        participant_ids = raw_df.loc[mask.fillna(False), "participant_id"]
        group_n = int(participant_ids.nunique())

        subset_df = long_df[long_df["participant_id"].isin(participant_ids)].copy()
        mean_df, anova_df = summarize_targets(subset_df)
        plot_rows: list[dict[str, object]] = []

        target_keys = ["target_kind", "scale", "item_no", "item_text", "score_type"]
        for key_values, target_df in subset_df.groupby(target_keys, dropna=False, observed=True):
            target_kind, scale, item_no, item_text, score_type = key_values
            target_anova = anova_df[
                (anova_df["target_kind"] == target_kind)
                & (anova_df["scale"] == scale)
                & (anova_df["score_type"] == score_type)
            ].copy()
            target_means = mean_df[
                (mean_df["target_kind"] == target_kind)
                & (mean_df["scale"] == scale)
                & (mean_df["score_type"] == score_type)
            ].copy()

            if pd.isna(item_no):
                target_anova = target_anova[target_anova["item_no"].isna()].copy()
                target_means = target_means[target_means["item_no"].isna()].copy()
            else:
                target_anova = target_anova[target_anova["item_no"] == item_no].copy()
                target_means = target_means[target_means["item_no"] == item_no].copy()

            if target_anova.empty or target_means.empty:
                continue

            slug = make_target_slug(str(target_kind), str(scale), item_no, str(score_type))
            plot_name = f"{slug}.png"
            create_interaction_plot(target_df.dropna(subset=["score"]).copy(), target_means, target_anova, plot_dir / plot_name)
            label_row = pd.Series(
                {
                    "target_kind": target_kind,
                    "scale": scale,
                    "item_no": item_no,
                    "score_type": score_type,
                }
            )
            plot_rows.append({"target": format_target_label(label_row), "file": f"plots/{plot_name}"})

        subset_df.to_csv(out_dir / "long_data.csv", index=False, encoding="utf-8-sig")
        mean_df.to_csv(out_dir / "condition_means.csv", index=False, encoding="utf-8-sig")
        anova_df.to_csv(out_dir / "anova_results.csv", index=False, encoding="utf-8-sig")
        write_summary(out_dir, stratum, group_n, overlap, mean_df, anova_df, plot_rows)

        stratum_index_rows.append(
            {
                "stratum_name": stratum["name"],
                "source_scale": stratum["source_scale"],
                "operator": stratum["operator_label"],
                "threshold": stratum["threshold"],
                "group_n": group_n,
                "threshold_equal_n": overlap,
                "summary_path": str(out_dir / "summary.md"),
                "anova_results_path": str(out_dir / "anova_results.csv"),
                "condition_means_path": str(out_dir / "condition_means.csv"),
                "long_data_path": str(out_dir / "long_data.csv"),
                "plots_dir": str(plot_dir),
            }
        )

    pd.DataFrame(stratum_index_rows).to_csv(OUTPUT_DIR / "strata_index.csv", index=False, encoding="utf-8-sig")
    print(OUTPUT_DIR / "strata_index.csv")


if __name__ == "__main__":
    main()

