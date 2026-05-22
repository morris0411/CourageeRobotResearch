from __future__ import annotations

import csv
import math
from pathlib import Path
from statistics import variance

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr, ttest_rel
from statsmodels.stats.anova import AnovaRM


ROOT = Path(__file__).resolve().parents[1]
ANALYSIS_DIR = Path(__file__).resolve().parent
MAPPING_PATH = ANALYSIS_DIR / "column_mapping_clean_data.csv"
EXCEL_PATH = next(ROOT.rglob("きれいデータ.xlsx"))

GENDER_COLUMN = "A"
GENDER_GROUPS = {
    1: {"label": "男性", "code": "male"},
    2: {"label": "女性", "code": "female"},
}

VIDEO_CONDITION_MAP = {
    "1": {"conflict": "葛藤あり", "action": "成功あり"},
    "2": {"conflict": "葛藤あり", "action": "成功なし"},
    "3": {"conflict": "葛藤なし", "action": "成功あり"},
    "4": {"conflict": "葛藤なし", "action": "成功なし"},
}
CONFLICT_ORDER = ["葛藤あり", "葛藤なし"]
ACTION_ORDER = ["成功あり", "成功なし"]

SUBJECTIVE_ITEM_TARGETS = {
    "1": "項目1",
    "2": "項目2",
    "3": "項目3",
}

ANOVA_EFFECT_LABELS = [
    ("conflict", "葛藤"),
    ("action", "成功"),
    ("conflict:action", "交互作用"),
]


def configure_plot_style() -> None:
    plt.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "MS Gothic", "sans-serif"]
    plt.rcParams["axes.unicode_minus"] = False


def load_mapping() -> list[dict[str, str]]:
    with MAPPING_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_sheet() -> pd.DataFrame:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    columns = [get_column_letter(i) for i in range(1, ws.max_column + 1)]
    df = pd.DataFrame(rows, columns=columns)
    df.insert(0, "participant_id", range(1, len(df) + 1))
    return df


def filter_by_gender(df: pd.DataFrame, gender_value: int) -> pd.DataFrame:
    gender_series = pd.to_numeric(df[GENDER_COLUMN], errors="coerce")
    filtered = df.loc[gender_series == gender_value].copy()
    filtered["participant_id"] = range(1, len(filtered) + 1)
    return filtered


def get_scale_columns(
    mapping_rows: list[dict[str, str]],
    scale: str,
    timing: str | None = None,
    video_no: str | None = None,
    item_no: str | None = None,
) -> list[dict[str, str]]:
    rows = []
    for row in mapping_rows:
        if row["scale"] != scale:
            continue
        if timing is not None and row["timing"] != timing:
            continue
        if video_no is not None and row["video_no"] != video_no:
            continue
        if item_no is not None and row["item_no_within_scale"] != item_no:
            continue
        rows.append(row)
    return sorted(rows, key=lambda r: int(r["excel_col_index"]))


def reverse_score(series: pd.Series) -> pd.Series:
    return 6 - series


def score_from_columns(df: pd.DataFrame, items: list[dict[str, str]]) -> pd.Series:
    cols = []
    for item in items:
        series = pd.to_numeric(df[item["excel_col_letter"]], errors="coerce")
        if item.get("reverse_scored") == "yes":
            series = reverse_score(series)
        cols.append(series)
    return pd.concat(cols, axis=1).mean(axis=1, skipna=False)


def run_anova(long_df: pd.DataFrame, dependent: str) -> pd.DataFrame:
    analysis_df = long_df[["participant_id", "conflict", "action", dependent]].dropna().copy()
    result = AnovaRM(
        data=analysis_df,
        depvar=dependent,
        subject="participant_id",
        within=["conflict", "action"],
    ).fit()
    table = result.anova_table.reset_index().rename(columns={"index": "effect"})
    table["partial_eta_sq"] = (
        table["F Value"] * table["Num DF"] / (table["F Value"] * table["Num DF"] + table["Den DF"])
    )
    return table


def p_to_stars(p_value: float) -> str:
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    return "ns"


def format_p_for_sentence(p_value: float) -> str:
    if p_value < 0.001:
        return "p < .001"
    return f"p = {p_value:.3f}".replace("0.", ".")


def safe_cohens_d_paired(x: pd.Series, y: pd.Series) -> float:
    diff = x - y
    std = diff.std(ddof=1)
    if pd.isna(std) or std == 0:
        return math.nan
    return float(diff.mean() / std)


def run_simple_effects(long_df: pd.DataFrame, dependent: str, interaction_p: float) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if interaction_p >= 0.05:
        return pd.DataFrame(
            columns=["comparison_type", "level", "group1", "group2", "t_value", "p_value", "cohens_d", "significance"]
        )

    for conflict_level in CONFLICT_ORDER:
        subset = long_df[long_df["conflict"] == conflict_level].pivot(
            index="participant_id", columns="action", values=dependent
        ).dropna()
        stat, p_value = ttest_rel(subset["成功あり"], subset["成功なし"])
        rows.append(
            {
                "comparison_type": "成功の単純主効果",
                "level": conflict_level,
                "group1": "成功あり",
                "group2": "成功なし",
                "t_value": stat,
                "p_value": p_value,
                "cohens_d": safe_cohens_d_paired(subset["成功あり"], subset["成功なし"]),
                "significance": p_to_stars(p_value),
            }
        )

    for action_level in ACTION_ORDER:
        subset = long_df[long_df["action"] == action_level].pivot(
            index="participant_id", columns="conflict", values=dependent
        ).dropna()
        stat, p_value = ttest_rel(subset["葛藤あり"], subset["葛藤なし"])
        rows.append(
            {
                "comparison_type": "葛藤の単純主効果",
                "level": action_level,
                "group1": "葛藤あり",
                "group2": "葛藤なし",
                "t_value": stat,
                "p_value": p_value,
                "cohens_d": safe_cohens_d_paired(subset["葛藤あり"], subset["葛藤なし"]),
                "significance": p_to_stars(p_value),
            }
        )

    return pd.DataFrame(rows)


def add_anova_text(ax: plt.Axes, anova_df: pd.DataFrame) -> None:
    lines = []
    for effect_key, label in ANOVA_EFFECT_LABELS:
        row = anova_df.loc[anova_df["effect"] == effect_key].iloc[0]
        lines.append(f"{label}: F={row['F Value']:.3f}, p={row['Pr > F']:.3f}")
    ax.text(
        0.02,
        0.98,
        "\n".join(lines),
        transform=ax.transAxes,
        ha="left",
        va="top",
        fontsize=10,
        bbox={"boxstyle": "round,pad=0.25", "facecolor": "white", "edgecolor": "none", "alpha": 0.9},
    )


def annotate_simple_effects(ax: plt.Axes, means_df: pd.DataFrame, simple_df: pd.DataFrame) -> None:
    if simple_df.empty:
        return

    x_positions = {"葛藤あり": 0, "葛藤なし": 1}
    ylim_top = ax.get_ylim()[1]
    ylim_bottom = ax.get_ylim()[0]
    y_span = ylim_top - ylim_bottom
    visible_idx = 0

    for row in simple_df.itertuples(index=False):
        if row.comparison_type != "成功の単純主効果" or row.p_value >= 0.05:
            continue
        x = x_positions[row.level]
        high = means_df.loc[means_df["conflict"] == row.level, "mean_score"].max()
        y = high + y_span * (0.05 + 0.06 * visible_idx)
        visible_idx += 1
        ax.plot([x - 0.08, x - 0.08, x + 0.08, x + 0.08], [y - 0.02 * y_span, y, y, y - 0.02 * y_span], color="black", lw=1)
        ax.text(x, y + 0.015 * y_span, row.significance, ha="center", va="bottom", fontsize=12)


def create_interaction_plot(
    long_df: pd.DataFrame,
    dependent: str,
    anova_df: pd.DataFrame,
    simple_df: pd.DataFrame,
    output_path: Path,
    y_limits: tuple[float, float],
    y_label: str,
) -> None:
    means_df = long_df.groupby(["conflict", "action"], observed=True)[dependent].mean().reset_index(name="mean_score")

    fig, ax = plt.subplots(figsize=(8.5, 5.2))
    x = np.arange(len(CONFLICT_ORDER))
    colors = {"成功あり": "#1f4e79", "成功なし": "#b03a2e"}

    for action_level in ACTION_ORDER:
        subset = means_df[means_df["action"] == action_level].set_index("conflict").loc[CONFLICT_ORDER].reset_index()
        ax.plot(x, subset["mean_score"], marker="o", linewidth=2.2, color=colors[action_level])
        y_end = subset["mean_score"].iloc[-1]
        offset = 0.10 if action_level == "成功あり" else -0.10
        ax.text(x[-1] + 0.08, y_end + offset, action_level, color=colors[action_level], fontsize=11, va="center")

    ax.set_xticks(x)
    ax.set_xticklabels(CONFLICT_ORDER)
    ax.set_ylabel(y_label)
    ax.set_xlabel("葛藤")
    ax.set_ylim(*y_limits)
    ax.set_xlim(-0.15, 1.55)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(False)

    add_anova_text(ax, anova_df)

    interaction_p = float(anova_df.loc[anova_df["effect"] == "conflict:action", "Pr > F"].iloc[0])
    if interaction_p < 0.05:
        annotate_simple_effects(ax, means_df, simple_df)

    fig.tight_layout()
    fig.savefig(output_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def create_scale_summary(
    output_path: Path,
    title: str,
    gender_label: str,
    n_participants: int,
    mean_table: pd.DataFrame,
    anova_tables: list[tuple[str, pd.DataFrame]],
    simple_tables: list[tuple[str, pd.DataFrame]],
    extra_sections: list[tuple[str, str]] | None = None,
    file_notes: list[str] | None = None,
) -> None:
    lines = [
        f"# {title}",
        "",
        f"- 元データ: `データ/{EXCEL_PATH.name}`",
        f"- 対象群: {gender_label}",
        f"- 対象人数: {n_participants}",
        "",
        "## 条件別平均",
        "",
        mean_table.to_markdown(index=False),
    ]

    for section_title, df in anova_tables:
        lines.extend(
            [
                "",
                f"## ANOVA: {section_title}",
                "",
                "| effect | F | Num DF | Den DF | p | partial η² |",
                "| --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for _, row in df.iterrows():
            lines.append(
                f"| {row['effect']} | {row['F Value']:.6f} | {row['Num DF']:.1f} | {row['Den DF']:.1f} | {row['Pr > F']:.6f} | {row['partial_eta_sq']:.6f} |"
            )

    for section_title, df in simple_tables:
        lines.extend(["", f"## 単純主効果: {section_title}", ""])
        if df.empty:
            lines.append("- 交互作用が有意でなかったため、単純主効果検定は実施せず")
        else:
            lines.append(df.to_markdown(index=False))

    if extra_sections:
        for section_title, body in extra_sections:
            lines.extend(["", f"## {section_title}", "", body])

    if file_notes:
        lines.extend(["", "## 出力ファイル", ""])
        lines.extend(file_notes)

    output_path.write_text("\n".join(lines), encoding="utf-8")


def cronbach_alpha(matrix: list[list[float]]) -> float:
    if not matrix or len(matrix) < 2:
        return math.nan
    item_count = len(matrix[0])
    if item_count < 2:
        return math.nan

    item_variances = [variance([row[col_idx] for row in matrix]) for col_idx in range(item_count)]
    total_scores = [sum(row) for row in matrix]
    total_variance = variance(total_scores)
    if total_variance == 0:
        return math.nan
    return (item_count / (item_count - 1)) * (1 - sum(item_variances) / total_variance)


def collect_complete_cases(
    raw_rows: list[dict[str, object]],
    items: list[dict[str, str]],
) -> tuple[list[list[float]], int]:
    matrix: list[list[float]] = []
    for raw_row in raw_rows:
        scored_row: list[float] = []
        valid = True
        for item in items:
            cell_value = raw_row[item["excel_col_letter"]]
            if cell_value is None or not isinstance(cell_value, (int, float)):
                valid = False
                break
            value = float(cell_value)
            if item["reverse_scored"] == "yes":
                value = 6 - value
            scored_row.append(value)
        if valid:
            matrix.append(scored_row)
    return matrix, len(matrix)


def format_alpha(alpha: float) -> str:
    if math.isnan(alpha):
        return ""
    return f"{alpha:.3f}"


def generate_internal_consistency(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    output_root: Path,
    gender_label: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    output_dir = output_root / "内的一貫性"
    output_dir.mkdir(parents=True, exist_ok=True)

    target_scales = {"主観の勇気評定", "勇気尺度", "CZO尺度", "葛藤尺度"}
    groups: dict[tuple[str, str, str, str], list[dict[str, str]]] = {}
    for row in mapping_rows:
        if row["scale"] not in target_scales:
            continue
        key = (row["section"], row["timing"], row["video_no"], row["scale"])
        groups.setdefault(key, []).append(row)

    raw_rows = df.drop(columns=["participant_id"]).to_dict("records")

    detailed_rows: list[dict[str, object]] = []
    for (section, timing, video_no, scale), items in sorted(groups.items(), key=lambda x: (x[0][1], x[0][2], x[0][3])):
        sorted_items = sorted(items, key=lambda r: int(r["excel_col_index"]))
        matrix, n_valid = collect_complete_cases(raw_rows, sorted_items)
        alpha = cronbach_alpha(matrix)
        reverse_items = [item["excel_col_letter"] for item in sorted_items if item["reverse_scored"] == "yes"]
        detailed_rows.append(
            {
                "section": section,
                "timing": timing,
                "video_no": video_no,
                "scale": scale,
                "n_items": len(sorted_items),
                "n_valid": n_valid,
                "cronbach_alpha": format_alpha(alpha),
                "item_columns": ", ".join(item["excel_col_letter"] for item in sorted_items),
                "item_numbers": ", ".join(item["item_no_within_scale"] for item in sorted_items),
                "reverse_scored_columns": ", ".join(reverse_items),
            }
        )

    pooled_rows: list[dict[str, object]] = []
    for scale in sorted(target_scales):
        post_groups = [
            sorted(
                get_scale_columns(mapping_rows, scale=scale, timing="事後", video_no=video_no),
                key=lambda r: int(r["excel_col_index"]),
            )
            for video_no in ["1", "2", "3", "4"]
            if get_scale_columns(mapping_rows, scale=scale, timing="事後", video_no=video_no)
        ]
        matrix: list[list[float]] = []
        for raw_row in raw_rows:
            for items in post_groups:
                scored_row: list[float] = []
                valid = True
                for item in items:
                    cell_value = raw_row[item["excel_col_letter"]]
                    if cell_value is None or not isinstance(cell_value, (int, float)):
                        valid = False
                        break
                    value = float(cell_value)
                    if item["reverse_scored"] == "yes":
                        value = 6 - value
                    scored_row.append(value)
                if valid:
                    matrix.append(scored_row)
        pooled_rows.append(
            {
                "section": "事後全条件プール",
                "timing": "事後",
                "video_no": "all",
                "scale": scale,
                "n_items": len(post_groups[0]) if post_groups else 0,
                "n_valid": len(matrix),
                "cronbach_alpha": format_alpha(cronbach_alpha(matrix)),
                "pooled_videos": len(post_groups),
            }
        )

    detailed_df = pd.DataFrame(detailed_rows)
    pooled_df = pd.DataFrame(pooled_rows)

    detailed_df.to_csv(output_dir / "cronbach_alpha_results.csv", index=False, encoding="utf-8-sig")
    pooled_df.to_csv(output_dir / "cronbach_alpha_results_pooled.csv", index=False, encoding="utf-8-sig")

    (output_dir / "cronbach_alpha_results.md").write_text(
        "\n".join(
            [
                "# 尺度の内的一貫性",
                "",
                f"- 対象群: {gender_label}",
                f"- 対象人数: {len(df)}",
                "",
                detailed_df.to_markdown(index=False),
            ]
        ),
        encoding="utf-8",
    )
    (output_dir / "cronbach_alpha_results_pooled.md").write_text(
        "\n".join(
            [
                "# 尺度の内的一貫性（事後全条件プール）",
                "",
                f"- 対象群: {gender_label}",
                f"- 対象人数: {len(df)}",
                "",
                pooled_df.to_markdown(index=False),
            ]
        ),
        encoding="utf-8",
    )

    return detailed_df, pooled_df


def generate_courage_analysis(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    output_root: Path,
    gender_label: str,
) -> dict[str, object]:
    output_dir = output_root / "ANOVA" / "勇気尺度"
    output_dir.mkdir(parents=True, exist_ok=True)

    pre_items = get_scale_columns(mapping_rows, "勇気尺度", timing="事前")
    pre_score = score_from_columns(df, pre_items)

    frames = []
    for video_no in ["1", "2", "3", "4"]:
        post_items = get_scale_columns(mapping_rows, "勇気尺度", timing="事後", video_no=video_no)
        post_score = score_from_columns(df, post_items)
        frames.append(
            pd.DataFrame(
                {
                    "participant_id": df["participant_id"],
                    "video_no": video_no,
                    "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                    "action": VIDEO_CONDITION_MAP[video_no]["action"],
                    "pre_score": pre_score,
                    "post_score": post_score,
                    "diff_score": post_score - pre_score,
                }
            )
        )

    long_df = pd.concat(frames, ignore_index=True)
    long_df["conflict"] = pd.Categorical(long_df["conflict"], categories=CONFLICT_ORDER, ordered=True)
    long_df["action"] = pd.Categorical(long_df["action"], categories=ACTION_ORDER, ordered=True)

    anova_post = run_anova(long_df, "post_score")
    anova_diff = run_anova(long_df, "diff_score")
    simple_post = run_simple_effects(
        long_df, "post_score", float(anova_post.loc[anova_post["effect"] == "conflict:action", "Pr > F"].iloc[0])
    )
    simple_diff = run_simple_effects(
        long_df, "diff_score", float(anova_diff.loc[anova_diff["effect"] == "conflict:action", "Pr > F"].iloc[0])
    )

    long_df.to_csv(output_dir / "courage_scores_long.csv", index=False, encoding="utf-8-sig")
    anova_post.to_csv(output_dir / "anova_post_score.csv", index=False, encoding="utf-8-sig")
    anova_diff.to_csv(output_dir / "anova_diff_score.csv", index=False, encoding="utf-8-sig")
    simple_post.to_csv(output_dir / "simple_effects_post_score.csv", index=False, encoding="utf-8-sig")
    simple_diff.to_csv(output_dir / "simple_effects_diff_score.csv", index=False, encoding="utf-8-sig")

    create_interaction_plot(
        long_df, "post_score", anova_post, simple_post, output_dir / "anova_plot_post_score.png", (1, 7), "勇気尺度得点"
    )
    create_interaction_plot(
        long_df, "diff_score", anova_diff, simple_diff, output_dir / "anova_plot_diff_score.png", (-1.0, 1.0), "勇気尺度差分"
    )

    mean_table = (
        long_df.groupby(["conflict", "action"], observed=True)[["post_score", "diff_score"]]
        .mean()
        .reset_index()
        .rename(columns={"post_score": "post_score_mean", "diff_score": "diff_score_mean"})
    )
    create_scale_summary(
        output_dir / "anova_summary.md",
        "勇気尺度の2要因分散分析",
        gender_label,
        len(df),
        mean_table,
        [("事後スコア", anova_post), ("差分スコア", anova_diff)],
        [("事後スコア", simple_post), ("差分スコア", simple_diff)],
        file_notes=[
            "- `courage_scores_long.csv`: 分析用ロングデータ",
            "- `anova_post_score.csv`: 事後スコアのANOVA表",
            "- `anova_diff_score.csv`: 差分スコアのANOVA表",
            "- `simple_effects_post_score.csv`: 事後スコアの単純主効果",
            "- `simple_effects_diff_score.csv`: 差分スコアの単純主効果",
            "- `anova_plot_post_score.png`: 事後スコアの交互作用プロット",
            "- `anova_plot_diff_score.png`: 差分スコアの交互作用プロット",
        ],
    )

    return {"post": anova_post, "diff": anova_diff, "means": mean_table}


def generate_czo_analysis(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    output_root: Path,
    gender_label: str,
) -> dict[str, object]:
    output_dir = output_root / "ANOVA" / "CZO尺度"
    output_dir.mkdir(parents=True, exist_ok=True)

    pre_items = get_scale_columns(mapping_rows, "CZO尺度", timing="事前")
    pre_score = score_from_columns(df, pre_items)

    frames = []
    for video_no in ["1", "2", "3", "4"]:
        post_items = get_scale_columns(mapping_rows, "CZO尺度", timing="事後", video_no=video_no)
        post_score = score_from_columns(df, post_items)
        frames.append(
            pd.DataFrame(
                {
                    "participant_id": df["participant_id"],
                    "video_no": video_no,
                    "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                    "action": VIDEO_CONDITION_MAP[video_no]["action"],
                    "pre_score": pre_score,
                    "post_score": post_score,
                    "diff_score": post_score - pre_score,
                }
            )
        )

    long_df = pd.concat(frames, ignore_index=True)
    long_df["conflict"] = pd.Categorical(long_df["conflict"], categories=CONFLICT_ORDER, ordered=True)
    long_df["action"] = pd.Categorical(long_df["action"], categories=ACTION_ORDER, ordered=True)

    anova_post = run_anova(long_df, "post_score")
    anova_diff = run_anova(long_df, "diff_score")
    simple_post = run_simple_effects(
        long_df, "post_score", float(anova_post.loc[anova_post["effect"] == "conflict:action", "Pr > F"].iloc[0])
    )
    simple_diff = run_simple_effects(
        long_df, "diff_score", float(anova_diff.loc[anova_diff["effect"] == "conflict:action", "Pr > F"].iloc[0])
    )

    long_df.to_csv(output_dir / "czo_scores_long.csv", index=False, encoding="utf-8-sig")
    anova_post.to_csv(output_dir / "anova_post_score.csv", index=False, encoding="utf-8-sig")
    anova_diff.to_csv(output_dir / "anova_diff_score.csv", index=False, encoding="utf-8-sig")
    simple_post.to_csv(output_dir / "simple_effects_post_score.csv", index=False, encoding="utf-8-sig")
    simple_diff.to_csv(output_dir / "simple_effects_diff_score.csv", index=False, encoding="utf-8-sig")

    create_interaction_plot(
        long_df, "post_score", anova_post, simple_post, output_dir / "anova_plot_post_score.png", (1, 5), "CZO尺度得点"
    )
    create_interaction_plot(
        long_df, "diff_score", anova_diff, simple_diff, output_dir / "anova_plot_diff_score.png", (-1.0, 1.0), "CZO尺度差分"
    )

    mean_table = (
        long_df.groupby(["conflict", "action"], observed=True)[["post_score", "diff_score"]]
        .mean()
        .reset_index()
        .rename(columns={"post_score": "post_score_mean", "diff_score": "diff_score_mean"})
    )
    create_scale_summary(
        output_dir / "anova_summary.md",
        "CZO尺度の2要因分散分析",
        gender_label,
        len(df),
        mean_table,
        [("事後スコア", anova_post), ("差分スコア", anova_diff)],
        [("事後スコア", simple_post), ("差分スコア", simple_diff)],
        file_notes=[
            "- `czo_scores_long.csv`: 分析用ロングデータ",
            "- `anova_post_score.csv`: 事後スコアのANOVA表",
            "- `anova_diff_score.csv`: 差分スコアのANOVA表",
            "- `simple_effects_post_score.csv`: 事後スコアの単純主効果",
            "- `simple_effects_diff_score.csv`: 差分スコアの単純主効果",
            "- `anova_plot_post_score.png`: 事後スコアの交互作用プロット",
            "- `anova_plot_diff_score.png`: 差分スコアの交互作用プロット",
        ],
    )

    return {"post": anova_post, "diff": anova_diff, "means": mean_table}


def generate_conflict_analysis(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    output_root: Path,
    gender_label: str,
) -> dict[str, object]:
    output_dir = output_root / "ANOVA" / "葛藤尺度"
    output_dir.mkdir(parents=True, exist_ok=True)

    frames = []
    for video_no in ["1", "2", "3", "4"]:
        scale_items = get_scale_columns(mapping_rows, "葛藤尺度", timing="事後", video_no=video_no)
        item5 = get_scale_columns(mapping_rows, "葛藤確認項目", timing="事後", video_no=video_no)[0]
        scale_score = score_from_columns(df, scale_items)
        item5_score = pd.to_numeric(df[item5["excel_col_letter"]], errors="coerce")
        frames.append(
            pd.DataFrame(
                {
                    "participant_id": df["participant_id"],
                    "video_no": video_no,
                    "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                    "action": VIDEO_CONDITION_MAP[video_no]["action"],
                    "scale_score": scale_score,
                    "item5_score": item5_score,
                }
            )
        )

    long_df = pd.concat(frames, ignore_index=True)
    long_df["conflict"] = pd.Categorical(long_df["conflict"], categories=CONFLICT_ORDER, ordered=True)
    long_df["action"] = pd.Categorical(long_df["action"], categories=ACTION_ORDER, ordered=True)

    anova_post = run_anova(long_df.rename(columns={"scale_score": "post_score"}), "post_score")
    simple_post = run_simple_effects(
        long_df.rename(columns={"scale_score": "post_score"}),
        "post_score",
        float(anova_post.loc[anova_post["effect"] == "conflict:action", "Pr > F"].iloc[0]),
    )

    corr_rows = []
    for video_no in ["1", "2", "3", "4"]:
        subset = long_df[long_df["video_no"] == video_no][["scale_score", "item5_score"]].dropna()
        r_value, p_value = pearsonr(subset["scale_score"], subset["item5_score"])
        corr_rows.append(
            {
                "scope": f"動画{video_no}",
                "video_no": video_no,
                "n": len(subset),
                "pearson_r": r_value,
                "p_value": p_value,
                "judgment_5pct": "有意" if p_value < 0.05 else "n.s.",
            }
        )
    pooled = long_df[["scale_score", "item5_score"]].dropna()
    r_value, p_value = pearsonr(pooled["scale_score"], pooled["item5_score"])
    corr_rows.append(
        {
            "scope": "全動画プール",
            "video_no": "all",
            "n": len(pooled),
            "pearson_r": r_value,
            "p_value": p_value,
            "judgment_5pct": "有意" if p_value < 0.05 else "n.s.",
        }
    )
    corr_df = pd.DataFrame(corr_rows)

    long_df.to_csv(output_dir / "conflict_scores_long.csv", index=False, encoding="utf-8-sig")
    anova_post.to_csv(output_dir / "anova_post_score.csv", index=False, encoding="utf-8-sig")
    simple_post.to_csv(output_dir / "simple_effects_post_score.csv", index=False, encoding="utf-8-sig")
    corr_df.to_csv(output_dir / "item5_correlation.csv", index=False, encoding="utf-8-sig")

    create_interaction_plot(
        long_df.rename(columns={"scale_score": "post_score"}),
        "post_score",
        anova_post,
        simple_post,
        output_dir / "anova_plot_post_score.png",
        (1, 7),
        "葛藤尺度得点",
    )

    mean_table = (
        long_df.groupby(["conflict", "action"], observed=True)[["scale_score"]]
        .mean()
        .reset_index()
        .rename(columns={"scale_score": "post_score_mean"})
    )
    create_scale_summary(
        output_dir / "anova_summary.md",
        "葛藤尺度の2要因分散分析と5項目目との相関",
        gender_label,
        len(df),
        mean_table,
        [("事後スコア", anova_post)],
        [("事後スコア", simple_post)],
        extra_sections=[("5項目目との相関", corr_df.to_markdown(index=False))],
        file_notes=[
            "- `conflict_scores_long.csv`: 分析用ロングデータ",
            "- `anova_post_score.csv`: 事後スコアのANOVA表",
            "- `simple_effects_post_score.csv`: 事後スコアの単純主効果",
            "- `item5_correlation.csv`: 5項目目との相関",
            "- `anova_plot_post_score.png`: 事後スコアの交互作用プロット",
        ],
    )

    return {"post": anova_post, "means": mean_table, "corr": corr_df}


def generate_subjective_analysis(
    df: pd.DataFrame,
    mapping_rows: list[dict[str, str]],
    output_root: Path,
    gender_label: str,
) -> dict[str, object]:
    output_dir = output_root / "ANOVA" / "主観勇気評定"
    output_dir.mkdir(parents=True, exist_ok=True)

    frames = []
    for item_no, target_name in SUBJECTIVE_ITEM_TARGETS.items():
        item_rows = get_scale_columns(mapping_rows, "主観の勇気評定", timing="事後", item_no=item_no)
        for row in item_rows:
            video_no = row["video_no"]
            frames.append(
                pd.DataFrame(
                    {
                        "participant_id": df["participant_id"],
                        "target_name": target_name,
                        "item_no": item_no,
                        "item_text": row["item_text"],
                        "video_no": video_no,
                        "conflict": VIDEO_CONDITION_MAP[video_no]["conflict"],
                        "action": VIDEO_CONDITION_MAP[video_no]["action"],
                        "score": pd.to_numeric(df[row["excel_col_letter"]], errors="coerce"),
                    }
                )
            )
    long_df = pd.concat(frames, ignore_index=True)
    long_df["conflict"] = pd.Categorical(long_df["conflict"], categories=CONFLICT_ORDER, ordered=True)
    long_df["action"] = pd.Categorical(long_df["action"], categories=ACTION_ORDER, ordered=True)
    long_df.to_csv(output_dir / "subjective_courage_long.csv", index=False, encoding="utf-8-sig")

    summary_lines = [
        "# 主観勇気評定の2要因分散分析",
        "",
        f"- 元データ: `データ/{EXCEL_PATH.name}`",
        f"- 対象群: {gender_label}",
        f"- 対象人数: {len(df)}",
    ]

    anova_results: dict[str, pd.DataFrame] = {}
    item_texts: dict[str, str] = {}
    for item_no in SUBJECTIVE_ITEM_TARGETS:
        item_df = long_df[long_df["item_no"] == item_no].copy()
        item_texts[item_no] = str(item_df["item_text"].iloc[0])
        anova_df = run_anova(item_df, "score")
        interaction_p = float(anova_df.loc[anova_df["effect"] == "conflict:action", "Pr > F"].iloc[0])
        simple_df = run_simple_effects(item_df, "score", interaction_p)

        anova_df.to_csv(output_dir / f"anova_item{item_no}.csv", index=False, encoding="utf-8-sig")
        simple_df.to_csv(output_dir / f"simple_effects_item{item_no}.csv", index=False, encoding="utf-8-sig")
        create_interaction_plot(
            item_df,
            "score",
            anova_df,
            simple_df,
            output_dir / f"anova_plot_item{item_no}.png",
            (1, 7),
            f"主観勇気評定 項目{item_no}",
        )
        anova_results[item_no] = anova_df

        summary_lines.extend(
            [
                "",
                f"## 項目{item_no}",
                "",
                item_texts[item_no],
                "",
                "| effect | F | Num DF | Den DF | p | partial η² |",
                "| --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for _, row in anova_df.iterrows():
            summary_lines.append(
                f"| {row['effect']} | {row['F Value']:.6f} | {row['Num DF']:.1f} | {row['Den DF']:.1f} | {row['Pr > F']:.6f} | {row['partial_eta_sq']:.6f} |"
            )
        summary_lines.append("")
        summary_lines.append(f"- `anova_item{item_no}.csv`: 項目{item_no}のANOVA表")
        summary_lines.append(f"- `simple_effects_item{item_no}.csv`: 項目{item_no}の単純主効果")
        summary_lines.append(f"- `anova_plot_item{item_no}.png`: 項目{item_no}の交互作用プロット")

    (output_dir / "anova_summary.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    return {"anova": anova_results, "item_texts": item_texts}


def effect_sentence(anova_df: pd.DataFrame, effect_key: str) -> str:
    row = anova_df.loc[anova_df["effect"] == effect_key].iloc[0]
    return (
        f"F({row['Num DF']:.0f}, {row['Den DF']:.0f}) = {row['F Value']:.3f}, "
        f"{format_p_for_sentence(float(row['Pr > F']))}, partial η² = {row['partial_eta_sq']:.3f}"
    )


def generate_overall_report(
    output_root: Path,
    gender_label: str,
    n_participants: int,
    pooled_alpha_df: pd.DataFrame,
    conflict_result: dict[str, object],
    courage_result: dict[str, object],
    czo_result: dict[str, object],
    subjective_result: dict[str, object],
) -> None:
    report_path = output_root / "総合結果レポート.md"

    pooled_lines = []
    for row in pooled_alpha_df.itertuples(index=False):
        pooled_lines.append(f"- {row.scale}: α = {row.cronbach_alpha or 'NA'}")

    report_lines = [
        "# 総合結果レポート",
        "",
        f"- 対象データ: `データ/{EXCEL_PATH.name}`",
        f"- 対象群: {gender_label}",
        f"- 対象人数: {n_participants}",
        "",
        "## 1. 尺度の内的一貫性",
        "",
        *pooled_lines,
        "",
        "## 2. マニプレーションチェック: 葛藤尺度",
        "",
        f"- 葛藤の主効果: {effect_sentence(conflict_result['post'], 'conflict')}",
        f"- 成功の主効果: {effect_sentence(conflict_result['post'], 'action')}",
        f"- 交互作用: {effect_sentence(conflict_result['post'], 'conflict:action')}",
        "",
        conflict_result["means"].to_markdown(index=False),
        "",
        "![葛藤尺度](./ANOVA/葛藤尺度/anova_plot_post_score.png)",
        "",
        "## 3. 勇気尺度",
        "",
        f"- 事後スコア 葛藤の主効果: {effect_sentence(courage_result['post'], 'conflict')}",
        f"- 事後スコア 成功の主効果: {effect_sentence(courage_result['post'], 'action')}",
        f"- 事後スコア 交互作用: {effect_sentence(courage_result['post'], 'conflict:action')}",
        f"- 差分スコア 葛藤の主効果: {effect_sentence(courage_result['diff'], 'conflict')}",
        f"- 差分スコア 成功の主効果: {effect_sentence(courage_result['diff'], 'action')}",
        f"- 差分スコア 交互作用: {effect_sentence(courage_result['diff'], 'conflict:action')}",
        "",
        courage_result["means"].to_markdown(index=False),
        "",
        "![勇気尺度 事後](./ANOVA/勇気尺度/anova_plot_post_score.png)",
        "",
        "![勇気尺度 差分](./ANOVA/勇気尺度/anova_plot_diff_score.png)",
        "",
        "## 4. CZO尺度",
        "",
        f"- 事後スコア 葛藤の主効果: {effect_sentence(czo_result['post'], 'conflict')}",
        f"- 事後スコア 成功の主効果: {effect_sentence(czo_result['post'], 'action')}",
        f"- 事後スコア 交互作用: {effect_sentence(czo_result['post'], 'conflict:action')}",
        f"- 差分スコア 葛藤の主効果: {effect_sentence(czo_result['diff'], 'conflict')}",
        f"- 差分スコア 成功の主効果: {effect_sentence(czo_result['diff'], 'action')}",
        f"- 差分スコア 交互作用: {effect_sentence(czo_result['diff'], 'conflict:action')}",
        "",
        czo_result["means"].to_markdown(index=False),
        "",
        "![CZO尺度 事後](./ANOVA/CZO尺度/anova_plot_post_score.png)",
        "",
        "![CZO尺度 差分](./ANOVA/CZO尺度/anova_plot_diff_score.png)",
        "",
        "## 5. 主観勇気評定",
    ]

    for item_no, item_text in subjective_result["item_texts"].items():
        anova_df = subjective_result["anova"][item_no]
        report_lines.extend(
            [
                "",
                f"### 項目{item_no}",
                "",
                item_text,
                "",
                f"- 葛藤の主効果: {effect_sentence(anova_df, 'conflict')}",
                f"- 成功の主効果: {effect_sentence(anova_df, 'action')}",
                f"- 交互作用: {effect_sentence(anova_df, 'conflict:action')}",
                "",
                f"![主観勇気評定 項目{item_no}](./ANOVA/主観勇気評定/anova_plot_item{item_no}.png)",
            ]
        )

    report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")


def generate_for_gender(mapping_rows: list[dict[str, str]], all_df: pd.DataFrame, gender_value: int) -> None:
    group_info = GENDER_GROUPS[gender_value]
    gender_label = group_info["label"]
    output_root = ANALYSIS_DIR / gender_label
    output_root.mkdir(parents=True, exist_ok=True)

    df = filter_by_gender(all_df, gender_value)
    _, pooled_alpha_df = generate_internal_consistency(df, mapping_rows, output_root, gender_label)
    conflict_result = generate_conflict_analysis(df, mapping_rows, output_root, gender_label)
    courage_result = generate_courage_analysis(df, mapping_rows, output_root, gender_label)
    czo_result = generate_czo_analysis(df, mapping_rows, output_root, gender_label)
    subjective_result = generate_subjective_analysis(df, mapping_rows, output_root, gender_label)
    generate_overall_report(
        output_root=output_root,
        gender_label=gender_label,
        n_participants=len(df),
        pooled_alpha_df=pooled_alpha_df,
        conflict_result=conflict_result,
        courage_result=courage_result,
        czo_result=czo_result,
        subjective_result=subjective_result,
    )

    print(f"[{gender_label}] n={len(df)}")
    print(output_root / "総合結果レポート.md")


def main() -> None:
    configure_plot_style()
    mapping_rows = load_mapping()
    all_df = load_sheet()
    for gender_value in sorted(GENDER_GROUPS):
        generate_for_gender(mapping_rows, all_df, gender_value)


if __name__ == "__main__":
    main()
