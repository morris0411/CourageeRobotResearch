from __future__ import annotations

import csv
import math
from pathlib import Path
from statistics import variance

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr


ROOT = Path(__file__).resolve().parent / "分析"
MAPPING_PATH = ROOT / "column_mapping_clean_data.csv"
EXCEL_PATH = Path(__file__).resolve().parent / "データ" / "きれいデータ.xlsx"
OUT_DIR = ROOT / "内的一貫性"
OUT_CSV = OUT_DIR / "subjective_item13_reliability.csv"
OUT_MD = OUT_DIR / "subjective_item13_reliability.md"


def load_mapping() -> list[dict[str, str]]:
    with MAPPING_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_rows() -> list[dict[str, object]]:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({get_column_letter(i): v for i, v in enumerate(row, start=1)})
    return rows


def cronbach_alpha(matrix: list[list[float]]) -> float:
    if not matrix or len(matrix) < 2:
        return math.nan
    k = len(matrix[0])
    if k < 2:
        return math.nan
    item_variances = [variance([row[col] for row in matrix]) for col in range(k)]
    total_scores = [sum(row) for row in matrix]
    total_variance = variance(total_scores)
    if total_variance == 0:
        return math.nan
    return (k / (k - 1)) * (1 - sum(item_variances) / total_variance)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    mapping = load_mapping()
    rows = load_rows()

    item_pairs: list[tuple[float, float]] = []
    video_columns: dict[str, list[str]] = {}
    for video_no in ["1", "2", "3", "4"]:
        selected = [
            row for row in mapping
            if row["scale"] == "主観の勇気評定"
            and row["timing"] == "事後"
            and row["video_no"] == video_no
            and row["item_no_within_scale"] in {"1", "3"}
        ]
        selected = sorted(selected, key=lambda r: int(r["item_no_within_scale"]))
        video_columns[video_no] = [row["excel_col_letter"] for row in selected]

    for raw_row in rows:
        for video_no in ["1", "2", "3", "4"]:
            cols = video_columns[video_no]
            values: list[float] = []
            valid = True
            for col in cols:
                cell = raw_row[col]
                if cell is None or not isinstance(cell, (int, float)):
                    valid = False
                    break
                values.append(float(cell))
            if valid:
                item_pairs.append((values[0], values[1]))

    matrix = [list(pair) for pair in item_pairs]
    alpha = cronbach_alpha(matrix)
    item1 = [pair[0] for pair in item_pairs]
    item3 = [pair[1] for pair in item_pairs]
    r_value, p_value = pearsonr(item1, item3)
    spearman_brown = (2 * r_value) / (1 + r_value) if (1 + r_value) != 0 else math.nan

    result = {
        "section": "事後全条件プール",
        "scale": "主観の勇気評定（項目1+3）",
        "item_numbers": "1, 3",
        "n_items": 2,
        "n_valid": len(item_pairs),
        "cronbach_alpha": f"{alpha:.6f}",
        "inter_item_r": f"{r_value:.6f}",
        "inter_item_p": f"{p_value:.6f}",
        "spearman_brown": f"{spearman_brown:.6f}",
    }

    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(result.keys()))
        writer.writeheader()
        writer.writerow(result)

    md = "\n".join(
        [
            "# 主観勇気評定 項目1+3 の内的一貫性",
            "",
            "- 対象: 事後の主観勇気評定の項目1と項目3",
            "- 算出方法: 4動画分をプールして完全回答のみで算出",
            "",
            "| 指標 | 値 |",
            "| --- | ---: |",
            f"| n | {len(item_pairs)} |",
            f"| Cronbach's α | {alpha:.3f} |",
            f"| 項目間相関 r | {r_value:.3f} |",
            f"| Spearman-Brown | {spearman_brown:.3f} |",
        ]
    ) + "\n"
    OUT_MD.write_text(md, encoding="utf-8")


if __name__ == "__main__":
    main()
